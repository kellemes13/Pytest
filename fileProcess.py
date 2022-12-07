import os
from openpyxl import load_workbook

#wb=Workbook()
for file in os.listdir('ori'):
    #file_name=os.path.join('ori',file)
    print(file)
   # print(type(file))
    file_name=os.path.join('ori',file)
    wb=load_workbook(file_name)
    ws=wb.active
    ws.title="test"
    ws['A1']="test for Excel"+ file[:-5]
    fix_file_name=os.path.join('proc',file[:-4]+'_fix.xlsx')
    wb.save(fix_file_name)

