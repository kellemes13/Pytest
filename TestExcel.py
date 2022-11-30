from openpyxl import Workbook
wb = Workbook()

# grab the active worksheet
ws = wb.active

# Data can be assigned directly to cells
ws['A1'] = 42

# Rows can also be appended
ws.append([1, 2, 3])
ws.append([4, 5, 6])

# Python types will automatically be converted
import datetime
ws['A5'] = datetime.datetime.now()
ws.append([7, 8, 9])
# Save the file
wb.save("sample1.xlsx")